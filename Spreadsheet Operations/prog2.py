from openpyxl import Workbook
from openpyxl.styles import Font

wb=Workbook()
sheet=wb.active
sheet.title="Language"
wb.create_sheet(title="Capital")
lang=["Kannada","Telugu","Tamil"]
state=["Karnataka","Telangana","Tamilnadu"]
Capital=["Bengaluru","Hyderabad","Chennai"]
Code=["KA",'TS','TN']
sheet.cell(row=1,column=1).value="state"
sheet.cell(row=1,column=2).value="Language"
sheet.cell(row=1,column=3).value="Code"
sheet.cell(row=1,column=4).value="Capital"
f1=Font(bold=True)
for row in sheet ['A1:C1']:
    for cell in row:
        cell.font=f1
for i in range(2,5):
    sheet.cell(row=i,column=1).value=state[i-2]
    sheet.cell(row=i,column=2).value=lang[i-2]
    sheet.cell(row=i,column=3).value=Code[i-2]
    sheet.cell(row=i,column=4).value=Capital[i-2]
wb.save("demo.xlsx")
sheet=wb["Capital"]
sheet.cell(row=1,column=1).value="state"
sheet.cell(row=1,column=2).value="Language"
sheet.cell(row=1,column=3).value="Code"
sheet.cell(row=1,column=4).value="Capital"
f1=Font(bold=True)
for row in sheet ['A1:C1']:
    for cell in row:
        cell.font=f1
for i in range(2,5):
    sheet.cell(row=i,column=1).value=state[i-2]
    sheet.cell(row=i,column=2).value=lang[i-2]
    sheet.cell(row=i,column=3).value=Code[i-2]
    sheet.cell(row=i,column=4).value=Capital[i-2]
wb.save("demo.xlsx")
SrchCode=input('Enter State Code for finding Capital:')
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if data==SrchCode:
        print("Corresponding Capital for Code",SrchCode,"is",sheet.cell(row=i,column=4).value)
sheet=wb['Language']
SrchCode=input('Enter State Code for finding language:')
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if data==SrchCode:
        print("Corresponding Language for Code",SrchCode,"is",sheet.cell(row=i,column=2).value)
